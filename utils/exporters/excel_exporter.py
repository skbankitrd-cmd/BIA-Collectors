import pandas as pd
import os
import uuid
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ExcelExporter:
    """原子化 Excel 導出器：全系統統一格式規範"""
    
    def __init__(self):
        # 統一儲存路徑
        self.output_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../../BIA-Web-Solution/server/public/generated_reports"))
        os.makedirs(self.output_dir, exist_ok=True)
        
        # 樣式定義：台灣金融業標準
        self.header_fill = PatternFill(start_color="E60012", end_color="E60012", fill_type="solid") # 台新紅
        self.header_font = Font(color="FFFFFF", bold=True, size=12)
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def export(self, data: list, report_title: str) -> str:
        """執行導出，回傳相對路徑"""
        if not data: return ""
        
        df = pd.DataFrame(data)
        file_name = f"{report_title}_{datetime.now().strftime('%Y%m%d_%H%M')}_{str(uuid.uuid4())[:4]}.xlsx"
        full_path = os.path.join(self.output_dir, file_name)
        
        with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='報表內容')
            
            # 套用統一視覺規範
            ws = writer.sheets['報表內容']
            for cell in ws[1]: # 標題列
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 自動欄寬與框線
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = self.thin_border
                    if cell.row > 1:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')

        return f"/public/generated_reports/{file_name}"
